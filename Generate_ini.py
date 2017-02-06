#!/usr/bin/env python3
#21dec16, hxy created

import argparse
from openpyxl import Workbook
from openpyxl import load_workbook
from collections import defaultdict

def get_l4_tcp_data(filename):
    """
    """
    dict_tcp = defaultdict(list)
    wb = load_workbook(filename)
    ws = wb.get_sheet_by_name('L4_Sender_via_iPerf')

    for row in ws.iter_rows(range_string='B9:D15', min_row=1, max_col=3, max_row=7):
        payload, ipv4, ipv6 = row
        dict_tcp[payload.value].append(ipv4.value)
        dict_tcp[payload.value].append(ipv6.value)
    return dict_tcp

def get_l4_udp_data(filename):
    """
    """
    dict_udp = defaultdict(list)
    wb = load_workbook(filename)
    ws = wb.get_sheet_by_name('L4_Sender_via_iPerf')

    for row in ws.iter_rows(range_string='B21:D24', min_row=1, max_col=3, max_row=4):
        payload, ipv4, ipv6 = row
        dict_udp[payload.value].append(ipv4.value)
        dict_udp[payload.value].append(ipv6.value)
    return dict_udp

def get_l3_v4_data(filename):
    """
    """
    dict_v4 = defaultdict(list)
    wb = load_workbook(filename)
    ws = wb.get_sheet_by_name('L3_IP_Forwarding')

    for row in ws.iter_rows(range_string='B9:D12', min_row=1, max_col=3, max_row=4):
        frame_size, flow_1, flow_2000 = row
        dict_v4[frame_size.value].append(flow_1.value)
        dict_v4[frame_size.value].append(flow_2000.value)
    return dict_v4

def get_l3_v6_data(filename):
    """
    """
    dict_v6 = defaultdict(list)
    wb = load_workbook(filename)
    ws = wb.get_sheet_by_name('L3_IP_Forwarding')

    for row in ws.iter_rows(range_string='B36:D39', min_row=1, max_col=3, max_row=4):
        frame_size, flow_1, flow_2000 = row
        dict_v6[frame_size.value].append(flow_1.value)
        dict_v6[frame_size.value].append(flow_2000.value)
    return dict_v6

def sort_data(dict_tcp, dict_udp, dict_v4, dict_v6):
    """
    """
    dict_list = []
    dict_list.extend(sorted(dict_tcp.items(), key=lambda d: d[0]))
    dict_list.extend(sorted(dict_udp.items(), key=lambda d: d[0]))
    dict_list.extend(sorted(dict_v4.items(), key=lambda d: d[0]))
    dict_list.extend(sorted(dict_v6.items(), key=lambda d: d[0]))
    return dict_list

def generate_ini_file(filename, dict_tcp, dict_udp, dict_v4, dict_v6):
    """
    """
    with open(filename.replace('.xlsx','_L4.ini'), 'wt') as f_l4, open(filename.replace('.xlsx','_L3.ini'), 'wt') as f_l3:
        print('[bsp]', file=f_l4)
        print('bsp=itl_64_vx7', file=f_l4)
        print('[mp]', file=f_l4)
        print('mp=SMP', file=f_l4)
        print('[bits]', file=f_l4)
        print('bits=64', file=f_l4)
        print('[tool]', file=f_l4)
        print('tool=gnu', file=f_l4)
        print('[layer]', file=f_l4)
        print('layer=L4', file=f_l4)

        print('[bsp]', file=f_l3)
        print('bsp=itl_64_vx7', file=f_l3)
        print('[mp]', file=f_l3)
        print('mp=SMP', file=f_l3)
        print('[bits]', file=f_l3)
        print('bits=64', file=f_l3)
        print('[tool]', file=f_l3)
        print('tool=gnu', file=f_l3)
        print('[layer]', file=f_l3)
        print('layer=L3', file=f_l3)

        case_num = 1
        for item in sort_data(dict_tcp, dict_udp, dict_v4, dict_v6):
            if case_num in list(range(1, 8)):
                print('[TestCase_%s]' %case_num, file=f_l4)
                print('testcase=tcp_%s' %item[0], file=f_l4)
                print('IPv4=%s' %item[1][0], file=f_l4)
                print('IPv6=%s' %item[1][1], file=f_l4)
            elif case_num in list(range(8,12)):
                print('[TestCase_%s]' %case_num, file=f_l4)
                print('testcase=udp_%s' %item[0], file=f_l4)
                print('IPv4=%s' %item[1][0], file=f_l4)
                print('IPv6=%s' %item[1][1], file=f_l4)
            elif case_num in list(range(12,16)):
                print('[TestCase_%s]' %(case_num-11), file=f_l3)
                print('testcase=IPv4_%s' %item[0], file=f_l3)
                print('1_flow=%s' %item[1][0], file=f_l3)
                print('2000_flow=%s' %item[1][1], file=f_l3)
            elif case_num in list(range(16,20)):
                print('[TestCase_%s]' %(case_num-11), file=f_l3)
                print('testcase=IPv6_%s' %item[0], file=f_l3)
                print('1_flow=%s' %item[1][0], file=f_l3)
                print('2000_flow=%s' %item[1][1], file=f_l3)
            case_num+=1


if __name__ == '__main__':
    parse = argparse.ArgumentParser()
    parse.add_argument('-f', '--file', help='xlsx file to process', dest='filename', required=True)
    args = parse.parse_args()
  
    filename = args.filename
    dict_tcp = get_l4_tcp_data(filename)
    dict_udp = get_l4_udp_data(filename)
    dict_v4 = get_l3_v4_data(filename)
    dict_v6 = get_l3_v6_data(filename)

    generate_ini_file(filename, dict_tcp, dict_udp, dict_v4, dict_v6)